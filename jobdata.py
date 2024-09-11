import logging
import os.path
import pickle
import sys
import time
from urllib.parse import parse_qs, urlparse

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


# url param 추출
def extract_param_value(url):
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    
    if "wantedAuthNo" in query_params:
        return query_params["wantedAuthNo"][0]
    else:
        return None

# 로깅 설정
logging.basicConfig(filename='my_log_file.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# login driver 설정
optionsLogin = webdriver.ChromeOptions()
optionsLogin.add_experimental_option('excludeSwitches', ['enable-logging'])
optionsLogin.add_experimental_option("excludeSwitches", ["enable-automation"])
optionsLogin.add_experimental_option("useAutomationExtension", False)

# ChromeOptions 및 WebDriverManager 설정
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_argument("headless")
options.add_argument('--log-level=3')
options.add_argument('--disable-loging')
options.add_experimental_option("useAutomationExtension", False)
service = Service(executable_path=ChromeDriverManager().install())

answer업종 = ""
answer고용허가제 = ""
answer직무내용 = ""
answer모집인원 = ""
answer근무지 = ""
answer공고일자 = ""
answer마감일자 = ""
answer임금조건 = ""
answer담당자이름 = ""
answer담당자전화번호 = ""
answer담당자휴대폰 = ""
answer담당자이메일 = ""
saveFlag = False
stopFlag = False

cookiefilelink = "C:\jobdata\worknetlogin.pkl"
firstfilelink = "C:\jobdata\\firstInfo.txt"
# 기타 변수 정의
excelPath = "C:\jobdata\jobData.xlsx"
excelEmailPath = "C:\jobdata\jobDataEmail.xlsx"

workbook = openpyxl.Workbook()
emailWorkbook = openpyxl.Workbook()


# 기존 엑셀 파일이 존재하는 경우 불러오기
if os.path.exists(excelPath):
    workbook = openpyxl.load_workbook(excelPath)

if os.path.exists(excelEmailPath):
    emailWorkbook = openpyxl.load_workbook(excelEmailPath)

# 엑셀 시트 및 컬럼 설정
sheet = workbook.active
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 50
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 50
sheet.column_dimensions['F'].width = 50
sheet.column_dimensions['G'].width = 50
sheet.column_dimensions['H'].width = 20
sheet.column_dimensions['I'].width = 20
sheet.column_dimensions['J'].width = 20
sheet.column_dimensions['K'].width = 20
sheet.column_dimensions['L'].width = 200
sheet.print_options.horizontalCentered = True
sheet.print_options.verticalCentered = True
sheet['A1'] = "업종"
sheet['B1'] = "직무내용"
sheet['C1'] = "모집인원"
sheet['D1'] = "근무지역"
sheet['E1'] = "공고일자"
sheet['F1'] = "마감일자"
sheet['G1'] = "임금조건"
sheet['H1'] = "담당자 이름"
sheet['I1'] = "담당자 전화번호"
sheet['J1'] = "담당자 휴대폰 번호"
sheet['K1'] = "수신자 Email 주소"
sheet['L1'] = "URL 공고 주소"
maxPagelen = 12

emailSheet = emailWorkbook.active
emailSheet.print_options.horizontalCentered = True
emailSheet.print_options.verticalCentered = True
emailSheet.column_dimensions['A'].width = 50
emailSheet['A1'] = "수신자 Email 주소"

currPage = input("검색 시작할 페이지 : ")
maxPagelen = input("검색 종료할 페이지 : ")
cloDateStdt = input("공고 마감 시작일 : ")
cloDateEndt = input("공고 마감 종료일 : ")
regDateStdt = input("공고 등록 시작일 : ")
regDateEndt = input("공고 등록 종료일 : ")

loginRedirectLink = ("https://www.work.go.kr/seekWantedMain.do")
loginMainLink = ("https://www.work24.go.kr/cm/z/b/0210/openLginPage.do?refSite=EAE05&refUrl=/g24Api/g24InterfaceSsoLogin.do?refUrl=/seekWantedMain.do")
mainLink = ("https://www.work.go.kr/empInfo/empInfoSrch/list/dtlEmpSrchList.do?"
        "careerTo=&keywordJobCd=&occupation=&templateInfo=&shsyWorkSecd=&rot2WorkYn=&payGbn=&resultCnt=10&keywordJobCont=N"
        f"&cert=&cloDateStdt={cloDateStdt}&moreCon=more&minPay=&codeDepth2Info=11000&isChkLocCall=&sortFieldInfo=DATE&major="
        "&resrDutyExcYn=&eodwYn=&sortField=DATE&staArea=&sortOrderBy=DESC&keyword=&termSearchGbn=all&carrEssYns="
        "&benefitSrchAndOr=O&disableEmpHopeGbn=&webIsOut=&actServExcYn=&maxPay=&keywordStaAreaNm=N&emailApplyYn="
        "&listCookieInfo=DTL&pageCode=&codeDepth1Info=11000&keywordEtcYn=&publDutyExcYn=&keywordJobCdSeqNo=&exJobsCd="
        f"&templateDepthNmInfo=&computerPreferential=&regDateStdt={regDateStdt}&employGbn=&empTpGbcd=&region=&infaYn=&resultCntInfo=10"
        f"&siteClcd=all&cloDateEndt={cloDateEndt}&sortOrderByInfo=DESC&currntPageNo=1&indArea=&careerTypes=&searchOn=Y&tlmgYn=&subEmpHopeYn="
        "&academicGbn=&templateDepthNoInfo=&foriegn=&mealOfferClcd=&station=&moerButtonYn=&holidayGbn=&srcKeyword="
        "&enterPriseGbn=all&academicGbnoEdu=noEdu&cloTermSearchGbn=all&keywordWantedTitle=N&stationNm=&benefitGbn="
        "&keywordFlag=&notSrcKeyword=&essCertChk=&isEmptyHeader=&depth2SelCode=&_csrf=dde2822b-952e-477d-81c3-17bb0c5d1775"
        f"&keywordBusiNm=N&preferentialGbn=&rot3WorkYn=&pfMatterPreferential=&regDateEndt={regDateEndt}&staAreaLineInfo1=11000"
        f"&staAreaLineInfo2=1&pageIndex={currPage}&termContractMmcnt=&careerFrom=&laborHrShortYn=#viewSPL")

loginDriver = webdriver.Chrome(service=service, options=optionsLogin)

try:
    loginDriver.get(loginMainLink)
    wait = WebDriverWait(loginDriver, 120)
    
    def check_url(loginDriver):
        return loginRedirectLink in loginDriver.current_url and loginMainLink not in loginDriver.current_url
        
        # 조건 함수를 사용하여 기다리기
    wait.until(check_url)
    
    print(loginDriver.current_url)
    
    if loginRedirectLink in loginDriver.current_url and loginDriver.current_url != loginMainLink :
        print("hello")
        pickle.dump(loginDriver.get_cookies(), open(cookiefilelink, "wb"))
            
except Exception as e:
    print(e)
    pass

finally:
    pickle.dump(loginDriver.get_cookies(), open(cookiefilelink, "wb"))
    loginDriver.close()



driver = webdriver.Chrome(service=service, options=options)
driver.get(mainLink)

driver.close()

if os.path.exists(firstfilelink):
    firstFile = open(firstfilelink, 'w')  # 'w' 모드는 파일을 비웁니다.
    firstFile.close()  # 파일을 비운 후에 닫습니다.

# 메인 링크 입력 받기
for curr in range(int(currPage)+1, int(maxPagelen) + 1):
    try:
        # 메인 WebDriver 생성 및 메인 링크 접속
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(mainLink)

        # 메인 페이지에서 joblinks 수집
        joblinks = driver.find_elements(By.CLASS_NAME, 'cp-info-in')

        # joblinks 반복 처리
        for linkss in joblinks:
            driverDetail = webdriver.Chrome(service=service, options=options)
            
            aTag = linkss.find_element(By.TAG_NAME, "a")
            aTagLink = aTag.get_attribute("href")

            wantedAuth = extract_param_value(aTagLink)+'\n'
            
            if os.path.exists(firstfilelink):
                firstFile = open(firstfilelink, 'r+')
                lines = firstFile.readlines()
                
                if wantedAuth in lines:
                    firstFile.write("duplicated Data \n")
                    print("중복된 데이터가 있습니다.")
                    stopFlag = True
                
                else:
                    firstFile.write(wantedAuth)
            else:
                firstFile = open(firstfilelink, 'a')
                firstFile.write(wantedAuth)
                    
            firstFile.close()
            
            if os.path.exists(cookiefilelink):
                workCookies = pickle.load(open(cookiefilelink, "rb"))
                driverDetail.get(loginRedirectLink)
                driverDetail.delete_all_cookies()
                
                for cookie in workCookies:
                    # cookie.pop("domain")
                    driverDetail.add_cookie(cookie)

                driverDetail.get(aTagLink)
            
            try:
                elements_careers_table = set(driverDetail.find_elements(By.CLASS_NAME, 'careers-table'))
                elements_업종_right = driverDetail.find_element(By.CLASS_NAME, 'right')
                elements_업종_li = elements_업종_right.find_elements(By.TAG_NAME, "li")
                elements_업종_strong = elements_업종_right.find_element(By.CLASS_NAME, "info").find_elements(By.TAG_NAME, "strong")

                # 업종 정보 수집
                for idx, val in enumerate(elements_업종_strong):
                    if val.text == "업종":
                        answer업종 = elements_업종_li.pop(idx).find_element(By.TAG_NAME, "div").text

                # 나머지 정보 수집
                for ele in elements_careers_table:
                    tableTh = ele.find_elements(By.TAG_NAME, "th")

                    for idx, val in enumerate(tableTh):
                        if val.text == "직무내용":
                            answers = ele.find_element(By.TAG_NAME, "td")
                            answer직무내용 = answers.text

                        if val.text == "모집인원":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer모집인원 = answers.pop(idx).text

                        if val.text == "근무예정지":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer근무지 = answers.pop(idx).text

                        if val.text == "임금조건":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer임금조건 = answers.pop(idx).text
                            
                        if val.text == "담당자":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer담당자이름 = answers.pop(idx).text
                            
                        if val.text == "전화번호":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer담당자전화번호 = answers.pop(idx).text
                            
                        if val.text == "휴대폰번호":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer담당자휴대폰 = answers.pop(idx).text
                            
                        if val.text == "이메일":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer담당자이메일 = answers.pop(idx).text
                            
                        if val.text == "채용공고 등록일시":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer공고일자 = answers.pop(idx).text
                            
                        if val.text == "접수마감일":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer마감일자 = answers.pop(idx).text

                        if val.text == "고용허가제":
                            answers = ele.find_elements(By.TAG_NAME, "td")
                            answer고용허가제 = answers.pop(idx).text

                            # 고용허가제가 있을 경우 데이터 저장
                            if answer고용허가제 != " ":
                                saveFlag = True
                                assertResult = aTagLink
                                print(assertResult)

                if saveFlag == True:
                    if stopFlag == False:
                        sheet.append([answer업종, answer직무내용, answer모집인원, answer근무지,answer공고일자, answer마감일자, answer임금조건,
                                                answer담당자이름, answer담당자전화번호, answer담당자휴대폰, answer담당자이메일, assertResult])
                        emailSheet.append([answer담당자이메일])
                        
                        emailWorkbook.save(excelEmailPath)
                        workbook.save(excelPath)
                        saveFlag = False
                    else:
                        stopFlag = False
                
            except NoSuchElementException as e:
                print(f"요소를 찾을 수 없습니다. 에러: {e.msg}")
                pass  # 요소를 찾을 수 없으면 패스 
            except Exception as e2:
                print(f"에러 요인 {e2}")
                pass             
            finally:
                stopFlag = False
                driverDetail.close()
        
        mainLink = mainLink.replace(f"pageIndex={curr-1}", f"pageIndex={curr}")
        driver.close()
        
    except Exception as e:
        print(f"에러 요인 {e}")
        pass
    
    finally:
        stopFlag = False
        time.sleep(1)
        print(f"현재 페이지 {curr-1} 입니다.")
        print(f"최대 페이지 {maxPagelen} 입니다.")
        if os.path.exists(firstfilelink):
            firstFile = open(firstfilelink, 'r+')
            firstFile.write(f"현재 페이지 {curr-1} 입니다.\n")    
        print("다음 페이지로 이동합니다.")
    
    # if stopFlag:
    #     break


# 엑셀 저장 및 WebDriver 종료
workbook.save(excelPath)
emailWorkbook.save(excelEmailPath)
print("종료")

driverDetail.quit()
loginDriver.quit()
driver.quit()